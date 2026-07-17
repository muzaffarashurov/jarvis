"""Business logic for orchestrating the Fast Response Board workbook.

Jarvis is not an Excel editor: FastResponseService never reads or
writes worksheet cells. It resolves configuration, checks file-system
facts about the workbook (existence, size, modification time), opens
the workbook through the existing ExecutionEngine (reusing
FileExecutor's OS-default-application behaviour, the same run() entry
point "system run <file>" already uses -- see SystemModule._run in
src/skills/system/skill.py), and creates plain file-system
backups with shutil.copy2.

Worksheet existence is checked by reading the workbook's sheet names
with openpyxl (already an existing project dependency, see
requirements.txt) in read-only mode -- this inspects workbook
structure only, never cell content.
"""

from __future__ import annotations

import os
import shutil
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path

from loguru import logger  # module-level logger, matching every other Jarvis component
from openpyxl import load_workbook

from src.core.command_router import CommandResult
from src.core.config import Config
from src.core.execution.engine import ExecutionEngine


class FastResponseConfigurationError(Exception):
    """Raised when 'fast_response.*' config entries are missing or invalid."""


@dataclass(frozen=True)
class WorkbookInfo:
    """Snapshot of the workbook used by both `frb status` and `frb info`.

    `last_modified` is timezone-aware UTC. `is_ready`/`error_message`
    back the "Ready / Error" line in `frb status`.
    """

    workbook_path: Path | None
    worksheet: str | None
    backup_folder: Path | None
    exists: bool
    size_bytes: int | None
    last_modified: datetime | None
    is_ready: bool
    error_message: str | None


@dataclass(frozen=True)
class ValidationResult:
    """Result of `frb validate`'s individual checks."""

    workbook_exists: bool
    worksheet_exists: bool
    workbook_readable: bool
    backup_folder_exists: bool

    @property
    def is_valid(self) -> bool:
        """Return True only if every individual check passed."""
        return (
            self.workbook_exists
            and self.worksheet_exists
            and self.workbook_readable
            and self.backup_folder_exists
        )


@dataclass(frozen=True)
class DoctorReport:
    """Result of `frb doctor`'s diagnostic checks."""

    configuration_loaded: bool
    workbook_exists: bool
    worksheet_configured: bool
    backup_directory_exists: bool
    permissions_ok: bool

    @property
    def is_ready(self) -> bool:
        """Return True only if every diagnostic check passed."""
        return (
            self.configuration_loaded
            and self.workbook_exists
            and self.worksheet_configured
            and self.backup_directory_exists
            and self.permissions_ok
        )


class FastResponseService:
    """Orchestrates the existing Fast Response Board Excel workbook.

    Responsibilities: resolve 'fast_response.*' configuration, report
    workbook facts, open the workbook via the shared ExecutionEngine,
    create timestamped backups, and validate readiness. Never reads or
    writes cell data; worksheet presence is checked purely by listing
    sheet names.
    """

    def __init__(self, config: Config, execution_engine: ExecutionEngine) -> None:
        """Initialize the FastResponseService.

        Args:
            config: Loaded application configuration.
            execution_engine: Shared engine used to open the workbook.
        """
        self._config = config
        self._execution_engine = execution_engine

    # ---------- Configuration resolution ----------

    def get_workbook_path(self) -> Path:
        """Resolve 'fast_response.workbook', raising if unconfigured."""
        return self._get_configured_path("fast_response.workbook")

    def get_backup_folder(self) -> Path:
        """Resolve 'fast_response.backup_folder', raising if unconfigured."""
        return self._get_configured_path("fast_response.backup_folder")

    def get_worksheet_name(self) -> str:
        """Resolve 'fast_response.worksheet', raising if unconfigured."""
        raw_value = self._config.get("fast_response.worksheet")
        if not isinstance(raw_value, str) or not raw_value.strip():
            raise FastResponseConfigurationError(
                "Missing or invalid 'fast_response.worksheet' entry in config/config.yaml."
            )
        return raw_value.strip()

    def _get_configured_path(self, key: str) -> Path:
        """Resolve a string config key to a Path, raising if invalid."""
        raw_value = self._config.get(key)
        if not isinstance(raw_value, str) or not raw_value.strip():
            raise FastResponseConfigurationError(
                f"Missing or invalid '{key}' entry in config/config.yaml."
            )
        return Path(raw_value.strip())

    def _resolve_all(self) -> tuple[Path, str, Path]:
        """Resolve workbook, worksheet, and backup folder together."""
        return self.get_workbook_path(), self.get_worksheet_name(), self.get_backup_folder()

    # ---------- Public API ----------

    def get_info(self) -> WorkbookInfo:
        """Return a full workbook snapshot for `frb status` / `frb info`."""
        try:
            workbook_path, worksheet, backup_folder = self._resolve_all()
        except FastResponseConfigurationError as exc:
            return WorkbookInfo(None, None, None, False, None, None, False, str(exc))

        if not workbook_path.is_file():
            return WorkbookInfo(
                workbook_path,
                worksheet,
                backup_folder,
                False,
                None,
                None,
                False,
                f"Workbook not found: {workbook_path}",
            )

        stat = workbook_path.stat()
        last_modified = datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc)
        return WorkbookInfo(
            workbook_path, worksheet, backup_folder, True, stat.st_size, last_modified, True, None
        )

    def open_workbook(self) -> CommandResult:
        """Open the workbook with the OS default application via ExecutionEngine."""
        try:
            workbook_path = self.get_workbook_path()
        except FastResponseConfigurationError as exc:
            logger.error(f"Fast Response Board open failed: {exc}")
            return CommandResult(success=False, message=str(exc))

        if not workbook_path.is_file():
            message = f"Workbook not found: {workbook_path}"
            logger.error(f"Fast Response Board open failed: {message}")
            return CommandResult(success=False, message=message)

        result = self._execution_engine.run(str(workbook_path))
        if not result.success:
            logger.error(f"Fast Response Board open failed: {result.message}")
            return CommandResult(
                success=False, message=f"Could not open workbook: {result.message}"
            )

        logger.info(f"Workbook opened: {workbook_path}")
        return CommandResult(success=True, message="Workbook opened.")

    def create_backup(self) -> CommandResult:
        """Create a timestamped copy of the workbook in the backup folder."""
        try:
            workbook_path = self.get_workbook_path()
            backup_folder = self.get_backup_folder()
        except FastResponseConfigurationError as exc:
            logger.error(f"Fast Response Board backup failed: {exc}")
            return CommandResult(success=False, message=str(exc))

        if not workbook_path.is_file():
            message = f"Workbook not found: {workbook_path}"
            logger.error(f"Fast Response Board backup failed: {message}")
            return CommandResult(success=False, message=message)

        if not backup_folder.is_dir():
            message = f"Backup folder not found: {backup_folder}"
            logger.error(f"Fast Response Board backup failed: {message}")
            return CommandResult(success=False, message=message)

        timestamp = datetime.now(timezone.utc).astimezone().strftime("%Y-%m-%d_%H%M%S")
        backup_name = f"{workbook_path.stem}_{timestamp}{workbook_path.suffix}"
        backup_path = backup_folder / backup_name

        try:
            shutil.copy2(workbook_path, backup_path)
        except OSError as exc:
            logger.error(f"Fast Response Board backup failed: {exc}")
            return CommandResult(success=False, message=f"Backup failed: {exc}")

        logger.info(f"Backup created: {backup_path}")
        return CommandResult(success=True, message=f"Backup created: {backup_name}")

    def validate(self) -> ValidationResult:
        """Run the `frb validate` checks against the workbook."""
        try:
            workbook_path, worksheet, backup_folder = self._resolve_all()
        except FastResponseConfigurationError:
            return ValidationResult(False, False, False, False)

        workbook_readable, worksheet_exists = self._check_workbook_readable(
            workbook_path, worksheet
        )
        return ValidationResult(
            workbook_exists=workbook_path.is_file(),
            worksheet_exists=worksheet_exists,
            workbook_readable=workbook_readable,
            backup_folder_exists=backup_folder.is_dir(),
        )

    def run_doctor(self) -> DoctorReport:
        """Run the `frb doctor` diagnostic checks."""
        try:
            workbook_path, worksheet, backup_folder = self._resolve_all()
        except FastResponseConfigurationError:
            return DoctorReport(False, False, False, False, False)

        workbook_exists = workbook_path.is_file()
        backup_directory_exists = backup_folder.is_dir()
        permissions_ok = (
            workbook_exists
            and os.access(workbook_path, os.R_OK)
            and backup_directory_exists
            and os.access(backup_folder, os.W_OK)
        )

        return DoctorReport(
            configuration_loaded=True,
            workbook_exists=workbook_exists,
            worksheet_configured=bool(worksheet),
            backup_directory_exists=backup_directory_exists,
            permissions_ok=permissions_ok,
        )

    # ---------- Internal helpers ----------

    @staticmethod
    def _check_workbook_readable(workbook_path: Path, worksheet: str) -> tuple[bool, bool]:
        """Return (workbook_readable, worksheet_exists) via openpyxl sheetnames."""
        if not workbook_path.is_file():
            return False, False

        try:
            workbook = load_workbook(filename=str(workbook_path), read_only=True)
            try:
                worksheet_exists = worksheet in workbook.sheetnames
            finally:
                workbook.close()
            return True, worksheet_exists
        except Exception as exc:  # noqa: BLE001 - a foreign/corrupted workbook can raise
            # many different exception types (openpyxl documents no exhaustive
            # hierarchy); this check's whole purpose is to report readability.
            logger.error(f"Fast Response Board workbook unreadable: {exc}")
            return False, False
